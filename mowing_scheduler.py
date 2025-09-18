#!/usr/bin/env python3
"""
Mowing Business Dynamic Scheduler - Fixed Version
Added proper file upload handling and settings page
"""

import os
import json
import io
import math
import sqlite3
import logging
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

import pandas as pd
import numpy as np
import requests

from flask import Flask, request, jsonify, send_file, render_template, redirect, url_for, flash
from werkzeug.utils import secure_filename
from plotly.utils import PlotlyJSONEncoder
import plotly.express as px
from openpyxl.styles import PatternFill, Font, Alignment

# Optional OR-Tools import guard
try:
    from ortools.constraint_solver import routing_enums_pb2, pywrapcp
    ORTOOLS_AVAILABLE = True
except Exception:
    ORTOOLS_AVAILABLE = False

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ====== Config ======
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"xlsx", "xls"}
DB_PATH = "mowing_scheduler.db"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ====== Enhanced ConfigManager ======
class ConfigManager:
    def __init__(self, file_path: str = "config.json"):
        self.file_path = file_path
        self.config = self._load()

    def _get_default_config(self) -> Dict[str, Any]:
        return {
            "business_settings": {
                "business_name": "ICC Mowing Services",
                "service_area": "Ipswich, QLD",
                "currency": "AUD",
                "timezone": "Australia/Brisbane"
            },
            "work_schedule": {
                "work_days_per_week": 5,
                "work_days": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
                "start_time": "06:30",
                "end_time": "15:00",
                "lunch_break_duration": 0.8,
                "include_weekends": False,
                "weekend_premium": 1.3
            },
            "team_defaults": {
                "max_daily_hours": 6.0,
                "overtime_allowed": True,
                "max_overtime_hours": 2.0,
                "base_hourly_rate": 29.0,
                "overtime_multiplier": 1.5,
                "default_mowing_rate_sqm_per_hour": 1000.0,
                "travel_time_same_suburb": 15,
                "travel_time_different_suburb": 30
            },
            "equipment_rates": {
                "commercial_mower": {
                    "sqm_per_hour": 1500,
                    "efficiency_factor": 1.0,
                    "maintenance_cost_per_hour": 5.0
                },
                "ride_on_mower": {
                    "sqm_per_hour": 800,
                    "efficiency_factor": 1.0,
                    "maintenance_cost_per_hour": 3.5
                },
                "zero_turn_mower": {
                    "sqm_per_hour": 1200,
                    "efficiency_factor": 0.9,
                    "maintenance_cost_per_hour": 6.0
                },
                "walk_behind_mower": {
                    "sqm_per_hour": 500,
                    "efficiency_factor": 1.2,
                    "maintenance_cost_per_hour": 2.0
                }
            },
            "difficulty_multipliers": {
                "1": 0.8,
                "2": 0.9,
                "3": 1.0,
                "4": 1.2,
                "5": 1.5
            },
            "weather_settings": {
                "api_key": "",
                "check_weather": True,
                "rain_threshold": 70,
                "wind_threshold": 25,
                "temperature_min": 5,
                "temperature_max": 35,
                "reschedule_on_bad_weather": True
            },
            "optimization_settings": {
                "max_parks_per_team_per_day": 8,
                "priority_weight": 2.0,
                "urgency_weight": 1.5,
                "travel_cost_weight": 1.0,
                "minimum_job_duration": 0.25,
                "setup_cleanup_time": 0.25
            },
            "reporting": {
                "include_costs": True,
                "include_weather": True,
                "include_maps": True,
                "export_format": "excel",
                "chart_colors": ["#667eea", "#764ba2", "#f093fb", "#f5576c", "#4facfe"]
            }
        }

    def _load(self) -> Dict[str, Any]:
        default = self._get_default_config()
        if os.path.exists(self.file_path):
            try:
                with open(self.file_path, "r") as f:
                    user = json.load(f)
                    # Deep merge
                    self._deep_merge(default, user)
                    logger.info(f"Loaded config from {self.file_path}")
            except Exception as e:
                logger.warning(f"Couldn't read config.json: {e}; using defaults.")
        else:
            # Create default config file
            try:
                with open(self.file_path, "w") as f:
                    json.dump(default, f, indent=2)
                logger.info(f"Created default config at {self.file_path}")
            except Exception as e:
                logger.error(f"Couldn't create config file: {e}")
        return default

    def _deep_merge(self, base: dict, override: dict):
        """Deep merge override into base"""
        for key, value in override.items():
            if key in base and isinstance(base[key], dict) and isinstance(value, dict):
                self._deep_merge(base[key], value)
            else:
                base[key] = value

    def get(self, path: str, default=None):
        keys = path.split(".")
        v = self.config
        for k in keys:
            if isinstance(v, dict) and k in v:
                v = v[k]
            else:
                return default
        return v

    def set(self, path: str, value):
        keys = path.split(".")
        d = self.config
        for k in keys[:-1]:
            if k not in d or not isinstance(d[k], dict):
                d[k] = {}
            d = d[k]
        d[keys[-1]] = value

    def save(self):
        try:
            with open(self.file_path, "w") as f:
                json.dump(self.config, f, indent=2)
            logger.info(f"Saved config to {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to save config: {e}")
            return False

    def update_from_form(self, form_data: dict):
        """Update config from web form data"""
        # Business Settings
        self.set("business_settings.business_name", form_data.get("business_name", ""))
        self.set("business_settings.service_area", form_data.get("service_area", ""))
        self.set("business_settings.currency", form_data.get("currency", "AUD"))
        
        # Work Schedule
        self.set("work_schedule.start_time", form_data.get("start_time", "06:30"))
        self.set("work_schedule.end_time", form_data.get("end_time", "15:00"))
        self.set("work_schedule.work_days_per_week", int(form_data.get("work_days_per_week", 5)))
        self.set("work_schedule.include_weekends", form_data.get("include_weekends") == "on")
        
        # Team Defaults
        self.set("team_defaults.max_daily_hours", float(form_data.get("max_daily_hours", 6.0)))
        self.set("team_defaults.base_hourly_rate", float(form_data.get("base_hourly_rate", 29.0)))
        self.set("team_defaults.overtime_allowed", form_data.get("overtime_allowed") == "on")
        self.set("team_defaults.max_overtime_hours", float(form_data.get("max_overtime_hours", 2.0)))
        self.set("team_defaults.default_mowing_rate_sqm_per_hour", float(form_data.get("default_mowing_rate_sqm_per_hour", 1000)))
        
        # Weather Settings
        self.set("weather_settings.api_key", form_data.get("weather_api_key", ""))
        self.set("weather_settings.check_weather", form_data.get("check_weather") == "on")
        self.set("weather_settings.rain_threshold", int(form_data.get("rain_threshold", 70)))
        
        return self.save()


config = ConfigManager()

# ====== Dataclasses ======
@dataclass
class Park:
    id: int
    name: str
    suburb: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    size_hectares: float = 0.0
    size_sqm: float = 0.0
    difficulty: int = 3
    last_mowed: Optional[datetime] = None
    priority: int = 3
    estimated_hours: float = 1.0


@dataclass
class Team:
    id: int
    name: str
    assigned_suburbs: List[str]
    max_daily_hours: float
    overtime_allowed: bool
    max_overtime_hours: float
    skills: List[str]
    equipment: List[str]
    hourly_rate: float
    mowing_rate_sqm_per_hour: float
    work_days: List[str]
    work_days_per_week: int


@dataclass
class ScheduledJob:
    team_id: int
    park_id: int
    start_time: datetime
    end_time: datetime
    is_overtime: bool
    travel_time: float
    estimated_cost: float


# ====== DB Manager (SQLite) ======
class DatabaseManager:
    def __init__(self, db_path: str = DB_PATH):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS parks (
                id INTEGER PRIMARY KEY,
                name TEXT,
                suburb TEXT,
                latitude REAL,
                longitude REAL,
                size_hectares REAL,
                size_sqm REAL,
                difficulty INTEGER,
                last_mowed TEXT,
                priority INTEGER,
                estimated_hours REAL
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS teams (
                id INTEGER PRIMARY KEY,
                name TEXT,
                assigned_suburbs TEXT,
                max_daily_hours REAL,
                overtime_allowed INTEGER,
                max_overtime_hours REAL,
                skills TEXT,
                equipment TEXT,
                hourly_rate REAL,
                mowing_rate_sqm_per_hour REAL,
                work_days TEXT,
                work_days_per_week INTEGER
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS scheduled_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                team_id INTEGER,
                park_id INTEGER,
                start_time TEXT,
                end_time TEXT,
                is_overtime INTEGER,
                travel_time REAL,
                estimated_cost REAL,
                created_at TEXT
            )
        ''')
        conn.commit()
        conn.close()

    def save_parks(self, parks: List[Park]):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        for p in parks:
            c.execute('''
                INSERT OR REPLACE INTO parks
                (id,name,suburb,latitude,longitude,size_hectares,size_sqm,difficulty,last_mowed,priority,estimated_hours)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (p.id, p.name, p.suburb, p.latitude, p.longitude,
                  p.size_hectares, p.size_sqm, p.difficulty,
                  p.last_mowed.isoformat() if p.last_mowed else None, p.priority, p.estimated_hours))
        conn.commit()
        conn.close()
        logger.info(f"Saved {len(parks)} parks to database")

    def load_parks(self) -> List[Park]:
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute('SELECT id,name,suburb,latitude,longitude,size_hectares,size_sqm,difficulty,last_mowed,priority,estimated_hours FROM parks')
        rows = c.fetchall()
        conn.close()
        parks = []
        for r in rows:
            parks.append(Park(
                id=r[0], name=r[1], suburb=r[2],
                latitude=r[3], longitude=r[4],
                size_hectares=r[5] or 0.0, size_sqm=r[6] or 0.0,
                difficulty=r[7] or 3,
                last_mowed=datetime.fromisoformat(r[8]) if r[8] else None,
                priority=r[9] or 3, estimated_hours=r[10] or 1.0
            ))
        return parks

    def save_teams(self, teams: List[Team]):
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        for t in teams:
            c.execute('''
                INSERT OR REPLACE INTO teams
                (id,name,assigned_suburbs,max_daily_hours,overtime_allowed,max_overtime_hours,skills,equipment,
                 hourly_rate,mowing_rate_sqm_per_hour,work_days,work_days_per_week)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (t.id, t.name, json.dumps(t.assigned_suburbs), t.max_daily_hours,
                  int(t.overtime_allowed), t.max_overtime_hours, json.dumps(t.skills),
                  json.dumps(t.equipment), t.hourly_rate, t.mowing_rate_sqm_per_hour,
                  json.dumps(t.work_days), t.work_days_per_week))
        conn.commit()
        conn.close()

    def load_teams(self) -> List[Team]:
        conn = sqlite3.connect(self.db_path)
        c = conn.cursor()
        c.execute('SELECT id,name,assigned_suburbs,max_daily_hours,overtime_allowed,max_overtime_hours,skills,equipment,hourly_rate,mowing_rate_sqm_per_hour,work_days,work_days_per_week FROM teams')
        rows = c.fetchall()
        conn.close()
        teams = []
        for r in rows:
            teams.append(Team(
                id=r[0], name=r[1],
                assigned_suburbs=json.loads(r[2]) if r[2] else [],
                max_daily_hours=r[3], overtime_allowed=bool(r[4]),
                max_overtime_hours=r[5], skills=json.loads(r[6]) if r[6] else [],
                equipment=json.loads(r[7]) if r[7] else [],
                hourly_rate=r[8], mowing_rate_sqm_per_hour=r[9],
                work_days=json.loads(r[10]) if r[10] else config.get("work_schedule.work_days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]),
                work_days_per_week=r[11] or 5
            ))
        return teams


# ====== Enhanced Excel Handler ======
class ExcelHandler:
    @staticmethod
    def import_parks_from_excel(path: str) -> List[Park]:
        try:
            logger.info(f"Importing parks from: {path}")
            df = pd.read_excel(path)
            logger.info(f"Excel file loaded with {len(df)} rows and columns: {list(df.columns)}")
            
            parks: List[Park] = []
            for idx, row in df.iterrows():
                try:
                    # Handle different column name variations
                    name = str(row.get('name', row.get('park_name', row.get('Name', f'Park {idx+1}'))))
                    suburb = str(row.get('suburb', row.get('Suburb', row.get('location', 'Unknown'))))
                    
                    # Coordinates
                    lat = None
                    lon = None
                    lat_cols = ['latitude', 'lat', 'Latitude', 'Lat']
                    lon_cols = ['longitude', 'lng', 'lon', 'Longitude', 'Lng', 'Lon']
                    
                    for col in lat_cols:
                        if col in row and not pd.isna(row[col]):
                            lat = float(row[col])
                            break
                    
                    for col in lon_cols:
                        if col in row and not pd.isna(row[col]):
                            lon = float(row[col])
                            break
                    
                    # Size handling
                    size_sqm = 0.0
                    size_ha = 0.0
                    
                    size_cols = ['size_sqm', 'area_sqm', 'sqm', 'Size_sqm', 'Area']
                    ha_cols = ['size_hectares', 'hectares', 'ha', 'Size_hectares']
                    
                    for col in size_cols:
                        if col in row and not pd.isna(row[col]):
                            size_sqm = float(row[col])
                            break
                    
                    for col in ha_cols:
                        if col in row and not pd.isna(row[col]):
                            size_ha = float(row[col])
                            break
                    
                    # Convert between hectares and sqm
                    if size_ha > 0 and size_sqm == 0:
                        size_sqm = size_ha * 10000
                    elif size_sqm > 0 and size_ha == 0:
                        size_ha = size_sqm / 10000
                    
                    # Default size if none provided
                    if size_sqm == 0 and size_ha == 0:
                        size_sqm = 5000.0  # Default 5000 sqm
                        size_ha = 0.5
                    
                    # Other fields with defaults
                    difficulty = int(row.get('difficulty', row.get('Difficulty', 3)))
                    priority = int(row.get('priority', row.get('Priority', 3)))
                    estimated_hours = float(row.get('estimated_hours', row.get('hours', 1.0)))
                    
                    # Last mowed date
                    last_mowed = None
                    mowed_cols = ['last_mowed', 'last_cut', 'previous_service']
                    for col in mowed_cols:
                        if col in row and not pd.isna(row[col]):
                            try:
                                last_mowed = pd.to_datetime(row[col])
                                break
                            except:
                                pass
                    
                    park = Park(
                        id=int(row.get('id', idx + 1)),
                        name=name,
                        suburb=suburb,
                        latitude=lat,
                        longitude=lon,
                        size_hectares=size_ha,
                        size_sqm=size_sqm,
                        difficulty=max(1, min(5, difficulty)),  # Clamp 1-5
                        last_mowed=last_mowed,
                        priority=max(1, min(5, priority)),  # Clamp 1-5
                        estimated_hours=max(0.1, estimated_hours)
                    )
                    parks.append(park)
                    
                except Exception as e:
                    logger.warning(f"Error processing row {idx}: {e}")
                    continue
            
            logger.info(f"Successfully imported {len(parks)} parks")
            return parks
            
        except Exception as e:
            logger.error(f"Excel import error: {e}")
            return []

    @staticmethod
    def export_schedule_to_excel(scheduled_jobs: List[ScheduledJob], parks: List[Park], teams: List[Team], out_path: str):
        try:
            park_map = {p.id: p for p in parks}
            team_map = {t.id: t for t in teams}
            rows = []
            for j in scheduled_jobs:
                p = park_map.get(j.park_id)
                t = team_map.get(j.team_id)
                rows.append({
                    "Team": t.name if t else f"Team {j.team_id}",
                    "Park": p.name if p else f"Park {j.park_id}",
                    "Suburb": p.suburb if p else "Unknown",
                    "Start": j.start_time.strftime("%Y-%m-%d %H:%M"),
                    "End": j.end_time.strftime("%Y-%m-%d %H:%M"),
                    "Duration_hours": round((j.end_time - j.start_time).total_seconds() / 3600, 2),
                    "Is_overtime": "Yes" if j.is_overtime else "No",
                    "Estimated_cost": j.estimated_cost,
                    "Travel_time_hours": j.travel_time
                })
            df = pd.DataFrame(rows)
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Schedule', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Schedule']
                # Simple header formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col in worksheet.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
            logger.info("Exported schedule to %s", out_path)
        except Exception as e:
            logger.error("Excel export error: %s", e)


# ====== Rest of classes remain the same ======
class DistanceCalculator:
    @staticmethod
    def haversine_distance(lat1, lon1, lat2, lon2):
        R = 6371.0
        lat1r, lon1r, lat2r, lon2r = map(math.radians, (lat1, lon1, lat2, lon2))
        dlat = lat2r - lat1r
        dlon = lon2r - lon1r
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1r) * math.cos(lat2r) * math.sin(dlon / 2) ** 2
        c = 2 * math.asin(math.sqrt(a))
        return R * c

    @staticmethod
    def estimate_travel_time_minutes(distance_km: float, avg_speed_kmh: float = 40.0) -> float:
        return (distance_km / avg_speed_kmh) * 60.0 if avg_speed_kmh > 0 else 30.0


class SchedulingOptimizer:
    def __init__(self, parks: List[Park], teams: List[Team]):
        self.parks = parks
        self.teams = teams
        self.distance_calc = DistanceCalculator()

    def calculate_mowing_time(self, park: Park, team: Team) -> float:
        rate = team.mowing_rate_sqm_per_hour or config.get("team_defaults.default_mowing_rate_sqm_per_hour", 1200)
        base = park.size_sqm / rate if rate > 0 else park.size_sqm / 1200
        setup = config.get("optimization_settings.setup_cleanup_time", 0.25)
        return max(config.get("optimization_settings.minimum_job_duration", 0.25), round(base + setup, 2))

    def is_work_day(self, date: datetime, team: Team) -> bool:
        return date.strftime("%A") in team.work_days

    def get_next_work_day(self, current_date: datetime, team: Team) -> datetime:
        next_date = current_date + timedelta(days=1)
        attempts = 0
        while not self.is_work_day(next_date, team) and attempts < 21:
            next_date += timedelta(days=1)
            attempts += 1
        return next_date

    def calculate_park_urgency(self, park: Park) -> float:
        if not park.last_mowed:
            return 1.0
        days = (datetime.now() - park.last_mowed).days
        if days <= 7:
            return 0.3
        elif days <= 14:
            return 0.6
        elif days <= 21:
            return 0.8
        else:
            return 1.0

    def optimize_schedule(self, start_date: datetime, num_days: int = 7) -> List[ScheduledJob]:
        scheduled: List[ScheduledJob] = []
        suburb_map: Dict[str, List[Park]] = {}
        for p in self.parks:
            suburb_map.setdefault(p.suburb, []).append(p)

        for team in self.teams:
            team_parks = []
            for s in team.assigned_suburbs:
                team_parks.extend(suburb_map.get(s, []))
            if not team_parks:
                continue
                
            team_parks.sort(key=lambda p: (-self.calculate_park_urgency(p), -p.priority, p.difficulty))
            current_date = start_date
            daily_hours = 0.0
            last_park = None

            if not self.is_work_day(current_date, team):
                current_date = self.get_next_work_day(current_date, team)

            for park in team_parks:
                duration = self.calculate_mowing_time(park, team)
                
                if last_park and last_park.latitude and park.latitude and last_park.longitude and park.longitude:
                    km = self.distance_calc.haversine_distance(last_park.latitude, last_park.longitude, park.latitude, park.longitude)
                    travel_mins = self.distance_calc.estimate_travel_time_minutes(km)
                else:
                    travel_mins = config.get("team_defaults.travel_time_same_suburb", 15) if park.suburb in team.assigned_suburbs else config.get("team_defaults.travel_time_different_suburb", 30)
                travel_hours = travel_mins / 60.0

                capacity = team.max_daily_hours + (team.max_overtime_hours if team.overtime_allowed else 0)
                if (daily_hours + duration + travel_hours) > capacity:
                    current_date = self.get_next_work_day(current_date, team)
                    daily_hours = 0.0
                    last_park = None

                is_overtime = (daily_hours + duration) > team.max_daily_hours

                base_cost = duration * team.hourly_rate
                overtime_cost = 0.0
                if is_overtime:
                    overtime_hours = max(0.0, (daily_hours + duration) - team.max_daily_hours)
                    overtime_cost = overtime_hours * team.hourly_rate * (config.get("team_defaults.overtime_multiplier", 1.5) - 1.0)

                maintenance = 0.0
                for eq in team.equipment:
                    rates = config.get("equipment_rates", {})
                    if eq in rates:
                        maintenance += rates[eq].get("maintenance_cost_per_hour", 0.0) * duration

                total_cost = round(base_cost + overtime_cost + maintenance, 2)

                start_of_day = datetime.combine(current_date.date(), datetime.strptime(config.get("work_schedule.start_time", "07:00"), "%H:%M").time())
                start_time = start_of_day + timedelta(hours=daily_hours)
                if last_park:
                    start_time += timedelta(hours=travel_hours)
                end_time = start_time + timedelta(hours=duration)

                scheduled.append(ScheduledJob(
                    team_id=team.id,
                    park_id=park.id,
                    start_time=start_time,
                    end_time=end_time,
                    is_overtime=is_overtime,
                    travel_time=round(travel_hours, 2),
                    estimated_cost=total_cost
                ))

                daily_hours += duration + travel_hours
                last_park = park

                if daily_hours >= capacity:
                    current_date = self.get_next_work_day(current_date, team)
                    daily_hours = 0.0
                    last_park = None

        return scheduled


class GanttChartGenerator:
    @staticmethod
    def create_gantt_json(scheduled_jobs: List[ScheduledJob], parks: List[Park], teams: List[Team]) -> str:
        try:
            park_map = {p.id: p for p in parks}
            team_map = {t.id: t for t in teams}
            rows = []
            for j in scheduled_jobs:
                p = park_map.get(j.park_id)
                t = team_map.get(j.team_id)
                rows.append({
                    "Task": f"{p.name if p else 'Park '+str(j.park_id)} ({p.suburb if p else 'Unknown'})",
                    "Start": j.start_time,
                    "Finish": j.end_time,
                    "Resource": t.name if t else f"Team {j.team_id}",
                    "Cost": j.estimated_cost
                })
            if not rows:
                return "{}"
            df = pd.DataFrame(rows)
            fig = px.timeline(df, x_start="Start", x_end="Finish", y="Resource", color="Cost", title="Mowing Schedule Gantt")
            fig.update_yaxes(autorange="reversed")
            return json.dumps(fig, cls=PlotlyJSONEncoder)
        except Exception as e:
            logger.error("Gantt creation error: %s", e)
            return "{}"


class ReportGenerator:
    @staticmethod
    def generate_summary(scheduled_jobs: List[ScheduledJob], parks: List[Park], teams: List[Team]) -> Dict[str, Any]:
        park_map = {p.id: p for p in parks}
        total_jobs = len(scheduled_jobs)
        total_cost = sum(j.estimated_cost for j in scheduled_jobs)
        total_hours = sum((j.end_time - j.start_time).total_seconds() / 3600 for j in scheduled_jobs)
        overtime_jobs = sum(1 for j in scheduled_jobs if j.is_overtime)
        team_stats = {}
        for t in teams:
            t_jobs = [j for j in scheduled_jobs if j.team_id == t.id]
            hours = sum((j.end_time - j.start_time).total_seconds() / 3600 for j in t_jobs)
            team_stats[t.name] = {
                "jobs": len(t_jobs),
                "hours": hours,
                "cost": sum(j.estimated_cost for j in t_jobs),
                "overtime_jobs": sum(1 for j in t_jobs if j.is_overtime),
                "utilization": (hours / (t.max_daily_hours * 7) * 100) if t.max_daily_hours > 0 else 0
            }
        suburb_stats = {}
        for p in parks:
            suburb_stats.setdefault(p.suburb, {"parks": 0, "jobs": 0, "cost": 0.0})
            suburb_stats[p.suburb]["parks"] += 1
            pj = [j for j in scheduled_jobs if j.park_id == p.id]
            suburb_stats[p.suburb]["jobs"] += len(pj)
            suburb_stats[p.suburb]["cost"] += sum(j.estimated_cost for j in pj)
        return {
            "overview": {"total_jobs": total_jobs, "total_cost": total_cost, "total_hours": total_hours, "overtime_jobs": overtime_jobs},
            "team_utilization": team_stats,
            "suburb_analysis": suburb_stats
        }


# ====== Defaults creators ======
def create_default_parks(n: int = 50) -> List[Park]:
    # Ipswich QLD suburbs
    suburbs = ['Ipswich Central', 'West Ipswich', 'East Ipswich', 'North Ipswich', 'Booval', 'Bundamba', 'Goodna', 'Redbank', 'Raceview', 'Yamanto', 'Leichhardt', 'One Mile', 'Silkstone', 'Brassall', 'Sadliers Crossing']
    parks = []
    for i in range(n):
        suburb = suburbs[i % len(suburbs)]
        # Ipswich coordinates: approximately -27.6149, 152.7594
        lat = -27.6149 + (i % 15) * 0.02 - 0.15 if np.random.random() > 0.3 else None
        lon = 152.7594 + (i % 15) * 0.02 - 0.15 if lat is not None else None
        size_sqm = float(np.random.uniform(5000, 25000))
        parks.append(Park(
            id=i + 1, name=f"Park {i+1}", suburb=suburb,
            latitude=lat, longitude=lon, size_hectares=size_sqm/10000.0, size_sqm=size_sqm,
            difficulty=int(np.random.randint(1, 6)), last_mowed=(datetime.now() - timedelta(days=int(np.random.randint(1, 30)))) if np.random.random() > 0.2 else None,
            priority=int(np.random.randint(1, 6)), estimated_hours=float(np.random.uniform(0.5, 4.0))
        ))
    return parks


def create_default_teams(num_teams: int = 13) -> List[Team]:
    # Ipswich QLD suburbs
    suburbs = ['Ipswich Central', 'West Ipswich', 'East Ipswich', 'North Ipswich', 'Booval', 'Bundamba', 'Goodna', 'Redbank', 'Raceview', 'Yamanto', 'Leichhardt', 'One Mile', 'Silkstone', 'Brassall', 'Sadliers Crossing']
    teams = []
    equip = list(config.get("equipment_rates", {}).keys()) or ['ride_on_mower']
    for i in range(num_teams):
        # Assign 2-3 suburbs per team
        start_idx = (i * 2) % len(suburbs)
        assigned = suburbs[start_idx:start_idx + 2] + ([suburbs[(start_idx + 2) % len(suburbs)]] if i % 3 == 0 else [])
        eq = np.random.choice(equip)
        mowing_rate = config.get("equipment_rates", {}).get(eq, {}).get("sqm_per_hour", config.get("team_defaults.default_mowing_rate_sqm_per_hour", 1200))
        mowing_rate += int(np.random.randint(-200, 200))
        teams.append(Team(
            id=i+1, name=f"Team {i+1}", assigned_suburbs=assigned,
            max_daily_hours=config.get("team_defaults.max_daily_hours", 8.0),
            overtime_allowed=config.get("team_defaults.overtime_allowed", True),
            max_overtime_hours=config.get("team_defaults.max_overtime_hours", 4.0),
            skills=['mowing', 'trimming'], equipment=[eq, 'trimmer'],
            hourly_rate=config.get("team_defaults.base_hourly_rate", 35.0) + float(np.random.uniform(-5, 10)),
            mowing_rate_sqm_per_hour=mowing_rate,
            work_days=config.get("work_schedule.work_days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]),
            work_days_per_week=5
        ))
    return teams


# ====== Flask App & Routes ======
app = Flask(__name__, template_folder="templates")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.secret_key = 'your-secret-key-here'  # For flash messages
db = DatabaseManager()

# Keep last results in memory for export/report
last_result: Dict[str, Any] = {}


@app.route("/")
def index():
    return render_template("dashboard.html", today=datetime.now().strftime("%Y-%m-%d"))


@app.route("/dashboard")
def dashboard():
    parks_count = len(db.load_parks())
    teams_count = len(db.load_teams())
    return render_template("dashboard.html", 
                         today=datetime.now().strftime("%Y-%m-%d"),
                         parks_count=parks_count,
                         teams_count=teams_count)


@app.route("/settings")
def settings_page():
    return render_template("settings.html", config=config.config)


@app.route("/map")
def map_view():
    return render_template("map.html")


@app.route("/upload_parks", methods=["POST"])
def upload_parks():
    try:
        if "file" not in request.files:
            flash("No file selected", "error")
            return redirect(url_for("dashboard"))
        
        file = request.files["file"]
        if file.filename == "":
            flash("No file selected", "error")
            return redirect(url_for("dashboard"))
        
        if not allowed_file(file.filename):
            flash("Invalid file type. Please upload Excel files (.xlsx, .xls) only", "error")
            return redirect(url_for("dashboard"))
        
        # Save file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)
        logger.info(f"File saved to: {filepath}")
        
        # Import parks
        parks = ExcelHandler.import_parks_from_excel(filepath)
        
        if parks:
            # Clear existing parks first (optional - remove if you want to append)
            db.save_parks(parks)  # This will replace all parks
            flash(f"Successfully imported {len(parks)} parks from {filename}", "success")
            logger.info(f"Successfully imported {len(parks)} parks")
            
            # Return JSON response if it's an AJAX request
            if request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
                return jsonify({"success": True, "parks_count": len(parks)})
        else:
            flash("No parks found in the uploaded file. Please check the format.", "error")
            if request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
                return jsonify({"success": False, "error": "No parks found in file"})
        
        # Cleanup uploaded file
        try:
            os.remove(filepath)
        except:
            pass
            
    except Exception as e:
        logger.error(f"Upload error: {e}")
        flash(f"Error uploading file: {str(e)}", "error")
        if request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
            return jsonify({"success": False, "error": str(e)})
    
    return redirect(url_for("dashboard"))


@app.route("/save_settings", methods=["POST"])
def save_settings():
    try:
        success = config.update_from_form(request.form)
        if success:
            flash("Settings saved successfully!", "success")
        else:
            flash("Error saving settings", "error")
    except Exception as e:
        logger.error(f"Settings save error: {e}")
        flash(f"Error saving settings: {str(e)}", "error")
    
    return redirect(url_for("settings_page"))


@app.route("/optimize", methods=["POST"])
def optimize_endpoint():
    try:
        num_teams = int(request.form.get("num_teams", 13))
        start_date_str = request.form.get("start_date", datetime.now().strftime("%Y-%m-%d"))
        duration = int(request.form.get("duration", 7))
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")

        parks = db.load_parks()
        if not parks:
            return jsonify({"success": False, "error": "No parks found. Please upload parks data first."}), 400

        teams = db.load_teams()
        if not teams or len(teams) != num_teams:
            teams = create_default_teams(num_teams)
            db.save_teams(teams)

        optimizer = SchedulingOptimizer(parks, teams)
        scheduled_jobs = optimizer.optimize_schedule(start_date, duration)

        global last_result
        last_result = {"scheduled_jobs": scheduled_jobs, "parks": parks, "teams": teams}

        stats = {
            "total_jobs": len(scheduled_jobs),
            "total_cost": round(sum(j.estimated_cost for j in scheduled_jobs), 2),
            "overtime_jobs": sum(1 for j in scheduled_jobs if j.is_overtime),
            "teams_used": len(set(j.team_id for j in scheduled_jobs))
        }

        gantt = GanttChartGenerator.create_gantt_json(scheduled_jobs, parks, teams)

        return jsonify({"success": True, "stats": stats, "gantt_chart": gantt})
    except Exception as e:
        logger.exception("Optimization failed")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/export")
def export_schedule():
    global last_result
    if not last_result:
        flash("No schedule available. Run optimization first.", "error")
        return redirect(url_for("dashboard"))
    
    try:
        out_path = os.path.join("/tmp", "mowing_schedule.xlsx")
        ExcelHandler.export_schedule_to_excel(last_result["scheduled_jobs"], last_result["parks"], last_result["teams"], out_path)
        return send_file(out_path, as_attachment=True, download_name="mowing_schedule.xlsx")
    except Exception as e:
        logger.error(f"Export error: {e}")
        flash(f"Export error: {str(e)}", "error")
        return redirect(url_for("dashboard"))


@app.route("/api/parks")
def api_parks():
    parks = db.load_parks()
    def serialize(p: Park):
        return {
            "id": p.id, "name": p.name, "suburb": p.suburb,
            "latitude": p.latitude, "longitude": p.longitude,
            "size_sqm": p.size_sqm, "difficulty": p.difficulty, "priority": p.priority
        }
    return jsonify([serialize(p) for p in parks])


@app.route("/api/teams")
def api_teams():
    teams = db.load_teams()
    return jsonify([asdict(t) for t in teams])


@app.route("/api/report")
def api_report():
    global last_result
    if not last_result:
        return jsonify({"error": "No optimization run yet"}), 400
    return jsonify(ReportGenerator.generate_summary(last_result["scheduled_jobs"], last_result["parks"], last_result["teams"]))


@app.route("/api/stats")
def api_stats():
    """Get current system statistics"""
    parks_count = len(db.load_parks())
    teams_count = len(db.load_teams())
    return jsonify({
        "parks_count": parks_count,
        "teams_count": teams_count
    })


@app.route("/api/config")
def api_config():
    return jsonify(config.config)


# ====== Run server ======
if __name__ == "__main__":
    logger.info("Starting Enhanced Mowing Scheduler Flask app")
    app.run(debug=True, host="0.0.0.0", port=5000)